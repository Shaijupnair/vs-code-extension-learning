from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Master Log"

# Define headers
headers = [
    "ID", "Start_Date", "End_Date", "Lead_Time_Days", "Category", 
    "Description", "Complexity_Points", "Active_Hours", "Is_Rework", 
    "Parent_ID", "Quality_Status"
]

# Add headers with formatting
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add example data
data = [
    ["TASK-001", datetime(2026, 1, 20), datetime(2026, 1, 22), None, "Coding", 
     "RAG Vectorization", 5, 4.0, "No", "", None],
    ["TASK-002", datetime(2026, 1, 25), datetime(2026, 1, 29), None, "Research", 
     "DAP Lifecycle", 3, 2.5, "No", "", None],
    ["TASK-003", datetime(2026, 2, 5), datetime(2026, 2, 5), None, "Rework", 
     "Fix memory bug", 0, 2.0, "Yes", "TASK-001", None]
]

for row_num, row_data in enumerate(data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=value)

# Add formulas for Lead_Time_Days (Column D)
for row in range(2, len(data) + 2):
    ws.cell(row=row, column=4, value=f"=C{row}-B{row}")

# Add formulas for Quality_Status (Column K)
for row in range(2, len(data) + 2):
    ws.cell(row=row, column=11, value=f'=IF(COUNTIF(J:J, A{row})>0, "⚠ Defective", "OK")')

# Add metrics section below the data
metrics_start_row = len(data) + 4

# Throughput Efficiency
ws.cell(row=metrics_start_row, column=1, value="Throughput Efficiency:")
ws.cell(row=metrics_start_row, column=1).font = Font(bold=True)
ws.cell(row=metrics_start_row, column=2, value="=SUM(G:G)/SUM(H:H)")
ws.cell(row=metrics_start_row + 1, column=1, value="(Points delivered per Active Hour)")
ws.cell(row=metrics_start_row + 1, column=1).font = Font(italic=True, size=9)

# Real Quality %
ws.cell(row=metrics_start_row + 3, column=1, value="Real Quality %:")
ws.cell(row=metrics_start_row + 3, column=1).font = Font(bold=True)
ws.cell(row=metrics_start_row + 3, column=2, value="=1 - (SUMIF(I:I, \"Yes\", H:H) / SUM(H:H))")
ws.cell(row=metrics_start_row + 4, column=1, value="(Percentage of time NOT spent on rework)")
ws.cell(row=metrics_start_row + 4, column=1).font = Font(italic=True, size=9)

# Format the metrics results as percentages
ws.cell(row=metrics_start_row + 3, column=2).number_format = '0.00%'

# Adjust column widths
column_widths = {
    'A': 12, 'B': 12, 'C': 12, 'D': 16, 'E': 12,
    'F': 25, 'G': 18, 'H': 14, 'I': 12, 'J': 12, 'K': 16
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save the workbook
filename = "Master_Log.xlsx"
wb.save(filename)
print(f"Excel file '{filename}' created successfully!")
print(f"\nThe file includes:")
print("- Headers with blue background")
print("- 3 example tasks with data")
print("- Lead_Time_Days formula (C-B)")
print("- Quality_Status formula (checks for defects)")
print("- Throughput Efficiency metric")
print("- Real Quality % metric")
